import json

from openpyxl import load_workbook

AttackType = ["恶意扫描", "远程命令", "XSS跨站", "SQL注入", "敏感文件访问",
              "Webshell", "代码执行", "恶意采集", "文件包含", "特殊攻击",
              "CC攻击", "其它"
              ]
City = ['河北', '辽宁', '陕西', '澳门', '天津', '江苏', '海南', '山东', '黑龙江',
        '台湾', '内蒙古', '河南', '福建', '宁夏', '青海', '上海', '四川', '湖北', '广东',
        '甘肃', '新疆', '北京', '浙江', '江西', '重庆', '广西', '山西', '西藏', '香港',
        '湖南', '云南', '吉林', '安徽', '贵州']

Project = "tsyy"


def UpdateExcel(date):
    x = 0
    y = 0
    now_col = 2  # 初始列为第2列
    excel_name = "D:/2020-云天/DataHandler/CYD/" + Project + "/report.xlsx"
    sheet_tt = (date[0] + "-" + date[-1])
    wb = load_workbook(excel_name)
    ws = wb.create_sheet(sheet_tt, 2)
    ws.append([""] + date + ["总计"])
    foreign_attack = {}
    foreign_numb = 0

    # 填充类型攻击名称
    for cell in ws["A2":"A13"]:
        cell[0].value = AttackType[x]
        x += 1

    ws["A14"].value = "动态阻断拦截"

    # 填充城市名称
    for cell in ws["A18":"A51"]:
        cell[0].value = City[y]
        y += 1

    for d in date:
        country_name = "D:/2020-云天/DataHandler/CYD/" + Project + "/CountryDate/" + d + ".txt"
        foreign_name = "D:/2020-云天/DataHandler/CYD/" + Project + "/ForeignDate/" + d + ".txt"
        attack_name = "D:/2020-云天/DataHandler/CYD/" + Project + "/AttackDate/" + d + ".txt"
        sum1 = 0
        sum2 = 0
        sum3 = 0
        sum4 = 0
        sum5 = 0
        sum6 = 0
        sum7 = 0
        sum8 = 0
        sum9 = 0
        sum10 = 0
        sum11 = 0
        sum12 = 0
        sum13 = 0

        # 统计单日各类型攻击数量
        with open(attack_name, encoding="UTF-8") as f:
            content = json.loads(f.read())
            f.close()
            for a in content["data"]["details"]:
                sum1 += a["base"]["恶意扫描"]
                sum2 += a["base"]["远程命令"]
                sum3 += a["base"]["XSS跨站"]
                sum4 += a["base"]["SQL注入"]
                sum5 += a["base"]["敏感文件访问"]
                sum6 += a["base"]["webshell"]
                sum7 += a["base"]["代码执行"]
                sum8 += a["base"]["恶意采集"]
                sum9 += a["base"]["文件包含"]
                sum10 += a["base"]["特殊攻击"]
                sum11 += a["base"]["CC攻击"]
                sum12 += a["base"]["其它"]
                sum13 += a["rule"]["动态阻断拦截"]

        # 统计国外地区攻击数量
        with open(foreign_name, encoding="UTF-8") as f:
            content = json.loads(f.read())
            for a in content["data"]["top10"]:
                if a["country"] in foreign_attack.keys():
                    foreign_attack[a["country"]] += a["waf_count"]
                else:
                    foreign_attack[a["country"]] = a["waf_count"]
                    foreign_numb += 1

        # 填充单日各类型攻击数量
        for col in ws.iter_cols(min_row=2, min_col=now_col, max_col=now_col, max_row=18):
            col[0].value = sum1
            col[1].value = sum2
            col[2].value = sum3
            col[3].value = sum4
            col[4].value = sum5
            col[5].value = sum6
            col[6].value = sum7
            col[7].value = sum8
            col[8].value = sum9
            col[9].value = sum10
            col[10].value = sum11
            col[11].value = sum12
            col[12].value = sum13
            col[13].value = f"=SUM({chr(now_col + 64)}2:{chr(now_col + 64)}13)"

        # 统计单日各地区攻击数量
        with open(country_name, encoding="UTF-8") as f:
            city_num = {}
            content = json.loads(f.read())
            f.close()
            for d in content["data"]:
                city_num[d["province"]] = d["waf_count"]

            for col in ws.iter_cols(min_row=18, min_col=now_col, max_col=now_col, max_row=17 + len(City)):
                z = 0
                for cell in col:
                    cell.value = city_num[City[z]]
                    z += 1

        now_col += 1

    # 填充各类型攻击总数
    z = 2
    for col in ws[f"{chr(now_col + 64)}2":f"{chr(now_col + 64)}15"]:
        for cell in col:
            cell.value = f"=SUM(B{z}:{chr(now_col + 63)}{z})"
            z += 1

    # 填充个城市攻击攻击总数
    z = 18
    for col in ws[f"{chr(now_col + 64)}18":f"{chr(now_col + 64)}51"]:
        for cell in col:
            cell.value = f"=SUM(B{z}:{chr(now_col + 63)}{z})"
            z += 1

    # 填充国外地区攻击数量/总数
    z = 0
    attack_sorted = sorted(foreign_attack.items(), reverse=True, key=lambda kv: (kv[1], kv[0],))
    for col in ws["A53":f"B{53 + foreign_numb - 1}"]:
        col[0].value = attack_sorted[z][0]
        col[1].value = attack_sorted[z][1]
        z += 1

    wb.save(excel_name)
    wb.close()
